"""Fill an AI-CAIQ from answers. The template is copied, never written.

Preserves rockin-robin's four maintenance rules verbatim, because they are the difference between
a questionnaire and a fiction:

1. **Every answer carries evidence.** Here that is stricter than rockin-robin's: evidence must be
   a ledger record id that verifies, not a sentence.
2. **"No" is legitimate.** A control we do not implement is answered No with the search path
   stated, not left blank to look better.
3. **Unassessed is not "No".** A control nobody looked at is left EMPTY. Writing No would claim a
   finding that was never made.
4. **Name the right control.** `IVS-*` does not exist in AICM v1.1; it is `I&S`. An `isalpha()`
   filter silently drops the ampersand in `A&A` and `I&S`.

Only columns C–F are written, per CSA's completion guidance, so the output still passes
rich-text's `verify_ai_caiq_workbook.py` unmodified.
"""

from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path

from stop_guessing.artifacts.digest import file_digest
from stop_guessing.caiq.workbook import (
    EXPECTED_SHEET,
    VALID_ANSWERS,
    VALID_OWNERS,
    WorkbookError,
    inspect,
)

COL_ANSWER, COL_OWNER, COL_IMPL, COL_CUSTOMER = 3, 4, 5, 6
COL_QUESTION_ID, COL_CONTROL_ID = 1, 9

CONTROL_RE = r"^[A-Z][A-Z&]{1,3}-\d{2}$"


class FillRefused(Exception):
    """Refuses rather than producing a plausible-looking workbook."""


@dataclass
class FillResult:
    output: Path
    template_digest_before: str
    template_digest_after: str
    controls_answered: int
    rows_written: int
    left_empty: int
    answers: dict[str, str]

    @property
    def template_untouched(self) -> bool:
        return self.template_digest_before == self.template_digest_after

    def to_dict(self) -> dict:
        return {
            "output": str(self.output),
            "output_digest": file_digest(self.output),
            "template_digest": self.template_digest_before,
            "template_untouched": self.template_untouched,
            "controls_answered": self.controls_answered,
            "rows_written": self.rows_written,
            "left_empty": self.left_empty,
        }


def _control_of(question_id: str) -> str:
    """`A&A-01.2` -> `A&A-01`. rsplit, because the ampersand must survive."""
    return question_id.rsplit(".", 1)[0]


def fill(template: str | Path, answers: dict[str, dict], output: str | Path) -> FillResult:
    """Copy the template, write only C:F, and verify the template is byte-identical afterwards.

    ``answers`` maps control id -> {answer, ssrm, implementation, customer_responsibilities}.
    """
    import openpyxl

    tpl, out = Path(template), Path(output)
    before = file_digest(tpl)
    if before is None:
        raise FillRefused(f"cannot read the template at {tpl}")

    ins = inspect(tpl)
    if ins.findings:
        raise FillRefused(
            "refusing to fill from a drifted template — regeneration is blocked until the "
            f"drift is resolved: {'; '.join(ins.findings)}"
        )

    for control, a in answers.items():
        if a.get("answer") not in VALID_ANSWERS:
            raise FillRefused(f"{control}: answer {a.get('answer')!r} not in {VALID_ANSWERS}")
        owner = a.get("ssrm")
        if a["answer"] == "NA" and owner:
            raise FillRefused(
                f"{control}: an NA answer must leave SSRM ownership blank (CSA guidance)"
            )
        if owner and owner not in VALID_OWNERS:
            raise FillRefused(f"{control}: ownership {owner!r} not in the CSA vocabulary")

    # Copy first. Nothing below ever opens the template for writing.
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(tpl, out)

    wb = openpyxl.load_workbook(out)
    try:
        if EXPECTED_SHEET not in wb.sheetnames:
            raise WorkbookError(f"{EXPECTED_SHEET} not in {wb.sheetnames}")
        ws = wb[EXPECTED_SHEET]

        seen: set[str] = set()
        rows_written = 0
        for row in range(3, ws.max_row + 1):
            qid = ws.cell(row, COL_QUESTION_ID).value
            if not isinstance(qid, str) or "." not in qid:
                continue
            control = _control_of(qid)
            a = answers.get(control)
            if a is None:
                continue  # unassessed is NOT "No" — leave it empty
            seen.add(control)
            ws.cell(row, COL_ANSWER).value = a["answer"]
            if a.get("ssrm"):
                ws.cell(row, COL_OWNER).value = a["ssrm"]
            if a.get("implementation"):
                ws.cell(row, COL_IMPL).value = a["implementation"]
            if a.get("customer_responsibilities"):
                ws.cell(row, COL_CUSTOMER).value = a["customer_responsibilities"]
            rows_written += 1

        unknown = set(answers) - seen
        if unknown:
            raise FillRefused(
                f"control id(s) not present in AICM v1.1.0: {sorted(unknown)}. "
                "Note IVS-* does not exist — it is I&S — and the ampersand in A&A/I&S is real."
            )
        wb.save(out)
    finally:
        wb.close()

    after = file_digest(tpl)
    if after != before:
        raise FillRefused(
            f"THE TEMPLATE WAS MODIFIED during fill ({before[:16]}… -> {str(after)[:16]}…). "
            "It is a copy-only local backup and is not recoverable from GitHub."
        )

    return FillResult(out, before, after, len(seen), rows_written,
                      left_empty=0, answers={k: v["answer"] for k, v in answers.items()})


def verify_with_rich_text(template: str | Path, filled: str | Path) -> tuple[bool, str]:
    """Run rich-text's `verify_ai_caiq_workbook.py` unmodified, if it is present.

    Deliberately an external check: our own verifier agreeing with our own filler proves only
    that we are self-consistent.
    """
    import os
    import subprocess
    import sys
    from glob import glob

    # #76 (SG-HARD-043): this hardcoded one maintainer's absolute plugin-cache path, so the
    # "unmodified third-party verifier" that CLAIM-15 and CLAIM-21 lean on was discoverable on
    # exactly one machine. Same defect as the hardcoded AI-CAIQ template, in the same subsystem,
    # missed when that one was fixed. Searched in order, most explicit first.
    candidates = [
        os.environ.get("STOP_GUESSING_CAIQ_VERIFIER"),
        # The plugin cache under whichever config dir is active, not a fixed home.
        str(Path(os.environ.get("CLAUDE_CONFIG_DIR") or Path.home() / ".claude")
            / "plugins/cache/rich-text/rich-text/*/skills/rich-text/scripts"
            / "verify_ai_caiq_workbook.py"),
        str(Path.home()
            / ".claude/plugins/cache/rich-text/rich-text/*/skills/rich-text/scripts"
            / "verify_ai_caiq_workbook.py"),
        str(Path.home() / "Software/rich-text/skills/rich-text/scripts"
            / "verify_ai_caiq_workbook.py"),
    ]
    matches: list[str] = []
    tried: list[str] = []
    for cand in candidates:
        if not cand:
            continue
        tried.append(cand)
        hits = sorted(glob(cand))
        if hits:
            matches = hits
            break
    if not matches:
        return False, (
            "rich-text's verify_ai_caiq_workbook.py was not found. It is a third-party verifier "
            "and is deliberately not vendored here — set $STOP_GUESSING_CAIQ_VERIFIER to its path, "
            "or install the rich-text plugin. Searched:\n  " + "\n  ".join(tried)
        )
    res = subprocess.run(  # noqa: S603
        [sys.executable, matches[-1], "--template", str(template), "--filled", str(filled)],
        capture_output=True, timeout=180,
    )
    out = (res.stdout + res.stderr).decode("utf-8", "replace").strip()
    return res.returncode == 0, out
