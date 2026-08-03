"""Read the AI-CAIQ / AICM workbooks. Never write them.

**The blank template is a copy-only local backup.** It is gitignored upstream in rockin-robin
("Public CSA reference materials are cited, never committed") and is therefore *not recoverable
from GitHub*. Every function here opens read-only and none of them saves. `make_blank_ai_caiq_
template.py` in rockin-robin is stale — it hardcodes ``SHEET = "AI-CAIQv1.0.2"`` and would
KeyError on a v1.1.0 workbook — and "fixing" it by regenerating a blank over the backup would
destroy the backup. Don't.

The version gate this module exists for: **cell A1 of the data sheet carries a machine-readable
JSON blob**, and nothing in the estate parses it. Every existing consumer hardcodes the sheet
name instead, which is why the stale script fails loudly on a version bump only *after* it has
already been trusted. Sheet names are a convention; A1 is the declaration.
"""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path

from stop_guessing.artifacts.digest import file_digest

# Pinned expectation. A bump here is a deliberate act, recorded in CHANGELOG.md.
EXPECTED_SPEC_NAME = "AI Controls Matrix"
EXPECTED_SPEC_VERSION = "1.1.0"
EXPECTED_CAIQ_VERSION = "1.1.0"
EXPECTED_SHEET = "AI-CAIQv1.1.0"
EXPECTED_DIMENSIONS = "A1:L324"

# CSA's completion guidance: only columns C-F are answerable.
EDITABLE_COLUMNS = (3, 4, 5, 6)
VALID_ANSWERS = ("Yes", "No", "NA")
VALID_OWNERS = (
    "Owned by CSP", "Owned by MP", "Owned by OSP", "Owned by AP", "Owned by AIC",
    "Shared across the supply chain", "Shared CSP-MP", "Shared MP-OSP",
    "Shared OSP-AP", "Shared AP-AIC", "Not Determined",
)


class WorkbookError(Exception):
    """Raised instead of guessing. A workbook we cannot read is not a workbook we may assume."""


@dataclass(frozen=True)
class Inspection:
    """The result of one every-run version inspection."""

    path: str
    digest: str | None
    a1_raw: str | None
    specification_name: str | None
    specification_version: str | None
    caiq_version: str | None
    sheet_names: list[str]
    data_sheet: str | None
    dimensions: str | None
    findings: list[str]

    @property
    def ok(self) -> bool:
        return not self.findings

    def to_dict(self) -> dict:
        return asdict(self)


def _load(path: str | Path):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment issue, not logic
        raise WorkbookError("openpyxl is required to inspect AI-CAIQ workbooks") from exc
    p = Path(path)
    if not p.is_file():
        raise WorkbookError(f"no workbook at {p}")
    # read_only=True is load-bearing, not an optimisation: it makes an accidental save impossible.
    return openpyxl.load_workbook(p, read_only=True, data_only=True)


def inspect(path: str | Path) -> Inspection:
    """Parse cell A1's JSON declaration and check it against the pinned expectation.

    Returns findings rather than raising, so a drifted workbook still produces a record. The
    caller decides what refuses — per the plan, drift blocks *regeneration and attestation*, never
    the session.
    """
    p = Path(path)
    findings: list[str] = []
    wb = _load(p)
    try:
        sheet_names = list(wb.sheetnames)
        data_sheet = EXPECTED_SHEET if EXPECTED_SHEET in sheet_names else None
        if data_sheet is None:
            findings.append(
                f"expected sheet {EXPECTED_SHEET!r} not present; sheets are {sheet_names}"
            )
            ws = wb[sheet_names[1]] if len(sheet_names) > 1 else wb[sheet_names[0]]
        else:
            ws = wb[data_sheet]

        # A read-only worksheet is "unsized" until forced; openpyxl raises rather than guessing,
        # which is the right call and means we have to ask explicitly.
        try:
            dimensions = ws.calculate_dimension(force=True)
        except (ValueError, TypeError) as exc:
            dimensions = None
            findings.append(f"cannot determine sheet dimensions: {exc}")
        a1 = ws.cell(row=1, column=1).value
        a1_raw = a1 if isinstance(a1, str) else None

        spec_name = spec_version = caiq_version = None
        if not a1_raw:
            findings.append("cell A1 is empty — no machine-readable version declaration")
        else:
            try:
                blob = json.loads(a1_raw)
            except json.JSONDecodeError as exc:
                findings.append(f"cell A1 is not JSON: {exc}")
            else:
                spec_name = blob.get("specification_name")
                spec_version = blob.get("specification_version")
                caiq_version = blob.get("caiq_version")
                if spec_name != EXPECTED_SPEC_NAME:
                    findings.append(
                        f"specification_name is {spec_name!r}, expected {EXPECTED_SPEC_NAME!r}"
                    )
                if spec_version != EXPECTED_SPEC_VERSION:
                    findings.append(
                        f"specification_version is {spec_version!r}, "
                        f"pinned to {EXPECTED_SPEC_VERSION!r}"
                    )
                if caiq_version is not None and caiq_version != EXPECTED_CAIQ_VERSION:
                    findings.append(
                        f"caiq_version is {caiq_version!r}, pinned to {EXPECTED_CAIQ_VERSION!r}"
                    )
        if dimensions and dimensions != EXPECTED_DIMENSIONS:
            findings.append(f"dimensions are {dimensions}, expected {EXPECTED_DIMENSIONS}")

        return Inspection(
            path=str(p),
            digest=file_digest(p),
            a1_raw=a1_raw,
            specification_name=spec_name,
            specification_version=spec_version,
            caiq_version=caiq_version,
            sheet_names=sheet_names,
            data_sheet=data_sheet,
            dimensions=dimensions,
            findings=findings,
        )
    finally:
        wb.close()


def template_record(path: str | Path) -> dict:
    """The committed `reference/TEMPLATE.json` — everything except the file itself.

    This is what makes the tool work for someone who does not have the blank template: the digest
    and the parsed declaration are pinned in git, so drift is still detectable against a template
    the operator obtains from CSA directly.
    """
    ins = inspect(path)
    return {
        "_note": (
            "Pinned facts about CSA's blank AI-CAIQ template. The template itself is COPY-ONLY "
            "and is not committed here — it is a local backup of a CSA artifact and is not "
            "recoverable from GitHub. Never modify or regenerate it. See IMPLEMENTATION_PLAN.md "
            "§12."
        ),
        "source_basename": Path(path).name,
        "sha256": ins.digest,
        "a1": ins.a1_raw,
        "specification_name": ins.specification_name,
        "specification_version": ins.specification_version,
        "caiq_version": ins.caiq_version,
        "sheet_names": ins.sheet_names,
        "data_sheet": ins.data_sheet,
        "dimensions": ins.dimensions,
        "editable_columns": list(EDITABLE_COLUMNS),
        "valid_answers": list(VALID_ANSWERS),
        "valid_owners": list(VALID_OWNERS),
    }
